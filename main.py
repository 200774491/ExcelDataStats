import tkinter as tk
from tkinter import filedialog
from functools import partial
import openpyxl
import os
import configparser

def export_selected_columns_to_txt(xlsm_file, sheet_names, output_txt, selected_column_headers, header_row=2, write_first_non_empty_only=True):
    try:
        wb = openpyxl.load_workbook(xlsm_file, read_only=True, keep_links=False)
        
        with open(output_txt, 'w', encoding='utf-8') as f:
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                headers = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
                headers = [header if header is not None else '' for header in headers]
                selected_columns = [idx for idx, header in enumerate(headers) if header in selected_column_headers]
                
                # 写入列标题
                selected_headers = [headers[idx] for idx in selected_columns]
                f.write(f"{sheet_name}\n")
                f.write('\t'.join(selected_headers) + '\n')
                
                if write_first_non_empty_only:
                    first_non_empty_values = {idx: None for idx in selected_columns}
                    
                    for row in sheet.iter_rows(min_row=header_row+1, values_only=True):
                        for col_idx in selected_columns:
                            if first_non_empty_values[col_idx] is None:
                                value = row[col_idx]
                                if value is not None:
                                    first_non_empty_values[col_idx] = str(value)
                    
                    # 写入所有列的第一个非空值
                    row_data = [first_non_empty_values[idx] if first_non_empty_values[idx] is not None else '' for idx in selected_columns]
                    f.write('\t'.join(row_data) + '\n')
                else:
                    # 写入列下的全部数据
                    for row in sheet.iter_rows(min_row=header_row+1, values_only=True):
                        row_data = [str(row[idx]) if row[idx] is not None else '' for idx in selected_columns]
                        f.write('\t'.join(row_data) + '\n')
        
        print(f"数据已成功导出到 {output_txt}")
    
    except FileNotFoundError:
        print(f"找不到文件：{xlsm_file}")
    except KeyError as e:
        print(f"找不到工作表：{str(e)}")
    except Exception as e:
        print(f"发生错误：{str(e)}")

def handle_export(xlsm_entry, sheets_entry, output_entry, headers_entry, non_empty_only_var):
    xlsm_file = xlsm_entry.get()
    sheet_names_str = sheets_entry.get()
    sheet_names = [name.strip() for name in sheet_names_str.split(',')]
    output_txt = output_entry.get()
    selected_headers_str = headers_entry.get()
    selected_column_headers = [header.strip() for header in selected_headers_str.split(',')]
    write_first_non_empty_only = non_empty_only_var.get()
    
    export_selected_columns_to_txt(xlsm_file, sheet_names, output_txt, selected_column_headers, header_row=2, write_first_non_empty_only=write_first_non_empty_only)

    # 保存配置到config.ini文件
    config = configparser.ConfigParser()
    config['Paths'] = {
        'ExcelFile': xlsm_file,
        'SheetNames': sheet_names_str,
        'OutputFile': output_txt,
        'SelectedHeaders': selected_headers_str,
        'WriteFirstNonEmptyOnly': str(write_first_non_empty_only)
    }
    with open('config.ini', 'w', encoding='utf-8') as configfile:
        config.write(configfile)

def browse_file(entry):
    filename = filedialog.askopenfilename()
    entry.delete(0, tk.END)
    entry.insert(0, filename)

def browse_output(entry):
    filename = filedialog.asksaveasfilename(defaultextension=".txt")
    entry.delete(0, tk.END)
    entry.insert(0, filename)

def load_config_to_ui(xlsm_entry, sheets_entry, output_entry, headers_entry, non_empty_only_var):
    if os.path.exists('config.ini'):
        config = configparser.ConfigParser()
        with open('config.ini', 'r', encoding='utf-8') as configfile:
            config.read_file(configfile)
        if 'Paths' in config:
            paths = config['Paths']
            xlsm_entry.insert(0, paths.get('ExcelFile', ''))
            sheets_entry.insert(0, paths.get('SheetNames', ''))
            output_entry.insert(0, paths.get('OutputFile', ''))
            headers_entry.insert(0, paths.get('SelectedHeaders', ''))
            non_empty_only_var.set(paths.getboolean('WriteFirstNonEmptyOnly', fallback=True))

def create_ui():
    root = tk.Tk()
    root.title("Excel导出工具")
    
    xlsm_label = tk.Label(root, text="Excel文件路径:")
    xlsm_label.grid(row=0, column=0)
    xlsm_entry = tk.Entry(root, width=50)
    xlsm_entry.grid(row=0, column=1)
    xlsm_browse_button = tk.Button(root, text="浏览...", command=partial(browse_file, xlsm_entry))
    xlsm_browse_button.grid(row=0, column=2)
    
    sheets_label = tk.Label(root, text="工作表名称 (逗号分隔):")
    sheets_label.grid(row=1, column=0)
    sheets_entry = tk.Entry(root, width=50)
    sheets_entry.grid(row=1, column=1)
    
    output_label = tk.Label(root, text="输出文件路径:")
    output_label.grid(row=2, column=0)
    output_entry = tk.Entry(root, width=50)
    output_entry.grid(row=2, column=1)
    output_browse_button = tk.Button(root, text="浏览...", command=partial(browse_output, output_entry))
    output_browse_button.grid(row=2, column=2)
    
    headers_label = tk.Label(root, text="列标题 (逗号分隔):")
    headers_label.grid(row=3, column=0)
    headers_entry = tk.Entry(root, width=50)
    headers_entry.grid(row=3, column=1)
    
    non_empty_only_var = tk.BooleanVar(value=True)
    non_empty_only_checkbutton = tk.Checkbutton(root, text="只写入所有列的第一个非空值", variable=non_empty_only_var)
    non_empty_only_checkbutton.grid(row=4, columnspan=3)
    
    export_button = tk.Button(root, text="导出", command=partial(handle_export, xlsm_entry, sheets_entry, output_entry, headers_entry, non_empty_only_var))
    export_button.grid(row=5, column=1)
    
    # 加载配置到UI输入框
    load_config_to_ui(xlsm_entry, sheets_entry, output_entry, headers_entry, non_empty_only_var)
    
    root.mainloop()

if __name__ == "__main__":
    create_ui()
